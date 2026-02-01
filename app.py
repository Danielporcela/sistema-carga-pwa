from flask import Flask, render_template, request
import sqlite3
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import io
import re
from datetime import datetime

app = Flask(__name__)

# ================= CONFIGURAÇÕES =================
COL_INI = "B"
COL_FIM = "T"

# linhas onde ficam os códigos dos setores
LINHAS_SETORES = [3, 31, 59]

# linhas onde ficam as cargas totais correspondentes
LINHAS_CARGA = {
    3: 24,
    31: 52,
    59: 80
}

DB_NAME = "dados_setores.db"
# =================================================


# ================= BANCO DE DADOS =================
def conectar_db():
    return sqlite3.connect(DB_NAME)


def criar_tabelas():
    conn = conectar_db()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS cargas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            setor TEXT,
            carga REAL,
            mes TEXT
        )
    """)

    conn.commit()
    conn.close()


criar_tabelas()
# =================================================


# ================= UTILITÁRIOS ===================
def normalizar(texto):
    if texto is None:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(texto).upper())


def carregar_planilha(arquivo):
    wb = load_workbook(io.BytesIO(arquivo), data_only=True)
    return wb.active
# =================================================


# ================= PROCESSAMENTO =================
def extrair_dados(ws):
    dados = []
    col_ini = column_index_from_string(COL_INI)
    col_fim = column_index_from_string(COL_FIM)

    for linha_setor in LINHAS_SETORES:
        linha_carga = LINHAS_CARGA[linha_setor]

        for col in range(col_ini, col_fim + 1):
            setor = ws.cell(linha_setor, col).value
            carga = ws.cell(linha_carga, col).value

            if setor and isinstance(carga, (int, float)):
                dados.append({
                    "setor": str(setor).strip(),
                    "carga": float(carga)
                })

    return dados


def salvar_no_banco(dados):
    conn = conectar_db()
    cur = conn.cursor()
    mes_atual = datetime.now().strftime("%Y-%m")

    for item in dados:
        cur.execute(
            "INSERT INTO cargas (setor, carga, mes) VALUES (?, ?, ?)",
            (item["setor"], item["carga"], mes_atual)
        )

    conn.commit()
    conn.close()


def ranking_db(ordem):
    conn = conectar_db()
    cur = conn.cursor()

    ordem_sql = "ASC" if ordem == "crescente" else "DESC"

    cur.execute(f"""
        SELECT setor, SUM(carga) as total
        FROM cargas
        GROUP BY setor
        ORDER BY total {ordem_sql}
    """)

    dados = cur.fetchall()
    conn.close()

    return [{"setor": d[0], "carga": d[1]} for d in dados]
# =================================================


# ================= ROTAS =========================
@app.route("/", methods=["GET", "POST"])
def index():
    resultado = None
    ranking = []
    mensagem = None
    ordem = request.form.get("ordem", "crescente")

    if request.method == "POST":
        arquivo = request.files.get("arquivo")
        setor_digitado = request.form.get("setor")

        if arquivo:
            ws = carregar_planilha(arquivo.read())
            dados = extrair_dados(ws)

            salvar_no_banco(dados)

            # Pesquisa por setor
            if setor_digitado:
                cod = normalizar(setor_digitado)
                for d in dados:
                    if normalizar(d["setor"]) == cod:
                        resultado = d
                        break

            ranking = ranking_db(ordem)
            mensagem = "Planilha processada com sucesso."

    return render_template(
        "index.html",
        resultado=resultado,
        ranking=ranking,
        ordem=ordem,
        mensagem=mensagem
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)

